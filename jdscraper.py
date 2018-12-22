from lxml import html
import time
import datetime
import sys
import requests
from random import choice
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://www.justdial.com/"
city_names = []
city_urls = []
vendor_names = []
vendor_urls = []
recordpos = 2
page_nof = 1
start_page = 1
end_page = 55
outfile = ""
wb = Workbook()


reqhead = {
"Host": "www.justdial.com",
    "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:62.0) Gecko/20100101 Firefox/62.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://www.justdial.com/",
    "DNT": "1",
    "Cookie": "", #put your cookie here!
    "Connection": "keep-alive",
    "Cache-Control": "max-age=0",
    "TE": "Trailers"
}

proxys = ["http://220.194.213.242:8080",
"http://50.93.200.83:1080",
"http://218.8.67.150:8998",
"http://171.211.184.234:8118",
"http://124.93.251.11:808",
"http://171.116.255.159:8118",
"http://47.88.11.13:8088",
"http://104.224.37.114:3128",
"http://112.193.123.46:8998",
"http://81.82.173.131:80",
"http://222.51.167.243:8118",
"http://101.4.136.34:8080",
"http://120.76.203.31:80",
"http://40.84.24.155:80",
"http://58.176.46.248:80",
"http://124.88.67.24:843",
"http://117.62.91.14:8998",
"http://119.9.91.186:8080",
"http://50.93.200.212:1080",
"http://123.206.126.146:80",
"http://47.89.51.239:8088",
]
def getProxyList():
	proxyList =[]
	for x in proxys:
		if len(x) > 5 :
			if x[:4] != 'http':
				proxyList.append('http://'+ x.split('\n')[0])
			else:
				proxyList.append(x.split('\n')[0])

	proxyList = list(set(proxyList))
	return proxyList


proxy = getProxyList()

def cities():

	page = requests.get(url,headers=reqhead,proxies={"http": "{}".format(choice(proxy))},timeout =3)
	page = page.text
	page_soup = BeautifulSoup(page,"lxml")
	for city in page_soup.findAll('div',{'class':'col-md-9 col-sm-9 col-xs-9 pl-50 pr-50 pt-10 foot-right col-md-offset-3 col-sm-offset-3 col-xs-offset-3'}):
		for city_name in city.findAll('li'):
			city_names.append(city_name.find('a').text.strip()	)
			city_urls.append(city_name.find('a').get('href'))


def categories(index):

	try:
		page = requests.get(city_urls[index],headers=reqhead,proxies={"http": "{}".format(choice(proxy))},timeout =3)
	except:
		categories()
	page = page.text
	page_soup = BeautifulSoup(page,"lxml")
	for vendors in page_soup.findAll('div',{'class':'col-md-9 col-sm-9 col-xs-9 pl-50 pr-50 pt-10 foot-right'}):
		for vendor in vendors.findAll('li'):
			vendor_names.append(vendor.find('a').text.strip()	)
			vendor_urls.append(vendor.find('a').get('href'))


def vendor_info(store_url,worksheet1):
	
	try:
		store_page = requests.get(store_url,headers=reqhead,proxies={"http": "{}".format(choice(proxy))},timeout =3)
		store_page = store_page.text
		store_soup = BeautifulSoup(store_page,"html.parser")
		store_name = store_soup.find('span',{'class':'fn'}).text.strip()
		worksheet1.cell( recordpos,  1, store_name)

		store_address = store_soup.find('span',{'id':'fulladdress'})
		if store_address is not None:
			store_address = store_address.text.strip()
		else:
			store_address = "NA"
		worksheet1.cell( recordpos,  2, store_address)

		store_email = store_soup.findAll('button',{'class':'jbtn fltrt'})
		try:
			store_email = store_email[1].get('onclick')
			a = store_email
			flag = 1
			b = ""
			for i in range(len(a)):
				if (a[i:i+6]) == "userid":
					for j in range(i+7,len(a)):
						if a[j] == "&":
							break
						b += a[j]
						flag = 0
				if flag == 0:
					break
			c = ""
			d = ""
			for i in range(len(b)):
				if b[i:i+3] == "%40":

					c = b[:i]
					d = b[i+4:]

			b = c + "@" + d
			if store_email is not None:
				store_email = b
				# print(store_email)
			else:
				store_email = "NA"
			if c == "" or d =="":
				store_email = "NA"

		except:
			store_email = "NA"
		worksheet1.cell( recordpos,  3, store_email)

		try:
			store_website = store_soup.findAll('span',{'class':'mreinfp comp-text'})
			store_website = store_website[len(store_website)-1].find('a').get('href')
		except:
			store_website = "NA"
		worksheet1.cell( recordpos,  4, store_website)
		store_categories = []

		try:
			store_whatspp = store_soup.find('a',{'id':'whatsapptriggeer'})
			store_whatspp_link = store_whatspp.get('href')
		except:
			store_whatspp_link = "NA"
		worksheet1.cell( recordpos,  5, store_whatspp_link)
		try:

			store_cate = store_soup.findAll('a',{'class':'lng_als_lst'})
			for cate in store_cate:
				store_categories.append(cate.text.strip())
		except:
			pass

		for cat in store_categories:
			worksheet1.cell( recordpos,  6+int(store_categories.index(cat)), cat)
		wb.save(outfile+".xlsx")
	except:
		vendor_info(store_url,worksheet1)


def scrape_vendor(worksheet1,url,start_page,end_page):
	global page_nof,outfile
	if int(page_nof)+ int(start_page)-1 > int(end_page):
		return
	try:
		page = requests.get(url,headers=reqhead,proxies={"http": "{}".format(choice(proxy))},timeout =3)
		print("Scraping page %s" %(page_nof+int(start_page)-1))
		page_nof += 1
		page = page.text	
		page_soup = BeautifulSoup(page,"lxml")
		global recordpos
		for store in page_soup.findAll('li',{'class':'cntanr'}):
			store_url = store.find('a').get('href')
			vendor_info(store_url,worksheet1)
			recordpos += 1	

			
		next_url = page_soup.find('a',{'rel':'next'})
		time.sleep(2)
		if next_url is not None:
			next_url = next_url.get('href')
			if next_url is not None:
				scrape_vendor(worksheet1,next_url,start_page,end_page)
	except:
		scrape_vendor(worksheet1,url,start_page,end_page)
	

def main():

	global store_page,end_page,outfile
	print("JustDial Scraper")
	print("################")
	print("################")
	outfile = input("Enter the name of output file\n")
	print("################")
	print("################")

	worksheet1 = wb.active
	recordpos = 1
	worksheet1.cell(1,1, 'Name')
	worksheet1.cell(1,2,'Address')
	worksheet1.cell(1,3,'Email')
	worksheet1.cell(1,4,'Website')
	worksheet1.cell(1,5,'Whatsapp Link')
	worksheet1.cell(1,6,'Categories')
	recordpos = 2	
	print("################")
	print("################")

	print("################")
	print("Enter 0 to enter the url and 1 to show all the cities")
	enter = input()
	if enter =="1":
		cities()

		print("Select the City to be scraped")
		for i in city_names:
			print(city_names.index(i),i)
		city_index = input()
		categories(int(city_index))
		print("################")
		print("################")
		print("################")
		print("Select the Vendor")
		for i in vendor_names:
			print(vendor_names.index(i),i)
		vendor_index = input()
		print("################")
		print("################")
		start_page = input("Enter the starting page number\n")
		end_page = input("Enter the ending page number\n")

		print("################")
		print("################")
		print("################")
		print("Scraping Started")
		scrape_vendor(worksheet1,vendor_urls[int(vendor_index)]+"/page-"+str(start_page),start_page,end_page)

	else:	
		new_url = input()
		start_page = input("Enter the starting page number\n")
		end_page = input("Enter the ending page number\n")
		scrape_vendor(worksheet1,new_url+"/page-"+str(start_page),start_page,end_page)
		
	wb.close()
	print("Scraping Completed")


if __name__ == "__main__":
	main()